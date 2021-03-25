from tkinter import *
from tkinter.ttk import *
from tkinter import messagebox
from openpyxl import *
from datetime import *
import backend
from autocompleteentry import AutocompleteEntry

# <editor-fold desc="Constants">
MONTH_KEYS = [None, '01', '02', '03', '04', '05',
              '06', '07', '08', '09', '10', '11', '12']
MONTH_NAMES = [None, 'January', 'February', 'March', 'April', 'May',
               'June', 'July', 'August', 'September', 'October', 'November', 'December']

MONTHS = {
    None: None, 
    '01': 'January', 
    '02': 'February', 
    '03': 'March', 
    '04': 'April', 
    '05': 'May',
    '06': 'June', 
    '07': 'July', 
    '08': 'August', 
    '09': 'September', 
    '10': 'October', 
    '11': 'November', 
    '12': 'December'
}

months = {}
for i, k in enumerate(MONTH_KEYS):
    months[k] = MONTH_NAMES[i]

months_inverse = {}
for i, k in enumerate(MONTH_NAMES):
    months_inverse[k] = MONTH_KEYS[i]

y, m, d = str(date.today()).split('-')
today_title = f'{months[m]} {d}, {y}'

mos = MONTH_NAMES[1:]
yrs = [i for i in range(2019, 2030)]
days = [i for i in range(1, 32)]
# </editor-fold>


class Main:
    data_title = 'R & C Trading Database - Malabuyoc'
    trans_title = 'R & C Trading - Malabuyoc'
    sales_title = 'No sales yet, try to update'
    data_dict = {}
    items = {}
    total = 0
    remain = 0
    due = 0

    data_list = []
    data = backend.view_data()
    for d in data:
        if d[1] not in data_list:
            data_list.append(d[1])

    def __init__(self):
        self.root = Tk()
        self.root.geometry(
            f'{self.root.winfo_screenwidth()}x{self.root.winfo_screenheight()}+0+0')
        self.root.resizable(False, False)
        self.root.title('R & C Trading - Malabuyoc')
        self.style = Style()
        self.style.configure('TNotebook.Tab', font='arial 20 italic')
        self.style.configure('TButton', font='arial 22 bold',
                             foreground='blue', width=9, relief=RAISED, borderwidth=5)
        self.style.configure('TLabel', font='arial 18 bold',
                             foreground='blue', padding=5)
        self.style.configure('Main.TFrame', background='cadet blue')
        self.style.configure('Treeview', font='arial 12 normal')
        self.tabs = Notebook(self.root, width=(
            self.root.winfo_screenwidth() - 40))
        self.tabs.pack()
        self.transac_win()
        self.data_win()
        self.sales_win()

    def data_win(self):

        # <editor-fold desc="Data Tab Functions">
        def add_new():
            data_keys = self.sort_data_dict()
            i, u, p, s = entry_item.get(), entry_unit.get(), entry_price.get(), entry_stock.get()
            if len(i) != 0 and len(u) != 0 and len(p) != 0 and len(s) != 0:
                key = f'{i}|{u}'
                if key not in data_keys:
                    try:
                        backend.add_record((i, u, float(p), float(s)))
                        self.update_data_list()
                    except ValueError:
                        messagebox.showwarning(
                            'Warning', '"Price" and "Stock" should contain numbers not words')
                else:
                    messagebox.showwarning(
                        'Warning', 'This item is already in database')
                search()
            else:
                messagebox.showwarning('Warning', "You can't add nothing")

        def display_data():
            data_keys = self.sort_data_dict()

            data = []
            for k in data_keys:
                i, u = k.split('|')
                p, s = self.data_dict[k][0], self.data_dict[k][1]
                d = [i, u, p, s]
                data.append(d)

            self.clear_tv(tv_data)
            self.display(tv_data, data)

        def clear_entries():
            entry_item.delete(0, END)
            entry_unit.delete(0, END)
            entry_price.delete(0, END)
            entry_stock.delete(0, END)
            entry_item.focus_set()

        def delete_data():
            i, u = entry_item.get(), entry_unit.get()
            if len(i) != 0 and len(u) != 0:
                k = backend.search_data(i, u)
                k = k[0][0]
                backend.delete_record(k)
                self.update_data_list()
                clear_entries()
                display_data()
            else:
                messagebox.showwarning(
                    'Warning', '"Item" and "Unit" should not be empty')

        def search():
            i, u = entry_item.get(), entry_unit.get()

            if len(i) != 0 or len(u) != 0:
                searched_data = backend.search_data(i, u)

                data = []
                for row in searched_data:
                    d = [row[1], row[2], row[3], row[4]]
                    data.append(d)

                self.clear_tv(tv_data)
                self.display(tv_data, data)

        def update():
            global k
            if isinstance(k, int):
                i, u, p, s = entry_item.get(), entry_unit.get(), entry_price.get(), entry_stock.get()
                backend.update_data(k, i, u, p, s)
                self.update_data_list()
                search()
            else:
                messagebox.showwarning('Warning', 'Select an item to update')

        def event_search(event):
            search()

        def tv_select(event):
            global k
            f = tv_data.focus()
            i = tv_data.item(f)
            x = i['values']
            k = backend.search_data(x[0], x[1])
            k = k[0][0]
            entry_item.delete(0, END)
            entry_item.insert(END, x[0])
            entry_unit.delete(0, END)
            entry_unit.insert(END, x[1])
            entry_price.delete(0, END)
            entry_price.insert(END, x[2])
            entry_stock.delete(0, END)
            entry_stock.insert(END, x[3])

        def event_item(event):
            entry_unit.focus_set()

        def event_unit(event):
            entry_price.focus_set()

        def event_price(event):
            entry_stock.focus_set()

        def event_stock(event):
            add_new()
        # </editor-fold>

        # <editor-fold desc="Data Frame">
        data_tab = Frame(self.tabs, style='Main.TFrame',
                         padding=(50, 0, 50, 0))
        data_tab.pack()

        # Add Data Tab
        self.tabs.add(data_tab, text='Database')

        # Data Tab Title
        self.title(data_tab, self.data_title)

        # <editor-fold desc="Data Mid Frame">
        d_mid_frame = Frame(data_tab, relief=SUNKEN,
                            borderwidth=5, padding=(30, 10, 30, 10))
        d_mid_frame.pack(fill=X)

        # <editor-fold desc="Data Mid Top Frame">
        d_mid_top = Frame(d_mid_frame)
        d_mid_top.pack(side=TOP, fill=X)

        # <editor-fold desc="Data Mid Top Right Frame">
        d_mid_top_right = Frame(d_mid_top, padding=(15, 65))
        d_mid_top_right.pack(side=RIGHT, fill=Y)

        # <editor-fold desc="Data Right Labels">
        lbl_item = Label(d_mid_top_right, text='Item', padding=(10, 20))
        lbl_unit = Label(d_mid_top_right, text='Unit', padding=(10, 20))
        lbl_price = Label(d_mid_top_right, text='Price', padding=(10, 20))
        lbl_stock = Label(d_mid_top_right, text='Stock', padding=(10, 20))

        lbl_item.grid(row=0, column=0, sticky=E)
        lbl_unit.grid(row=1, column=0, sticky=E)
        lbl_price.grid(row=2, column=0, sticky=E)
        lbl_stock.grid(row=3, column=0, sticky=E)
        # </editor-fold>

        # <editor-fold desc="Data Right Entries">
        entry_item = AutocompleteEntry(
            self.data_list, master=d_mid_top_right, font='arial 18 bold', foreground='green')
        entry_unit = Entry(
            d_mid_top_right, font='arial 18 bold', foreground='green')
        entry_price = Entry(
            d_mid_top_right, font='arial 18 bold', foreground='green')
        entry_stock = Entry(
            d_mid_top_right, font='arial 18 bold', foreground='green')

        entry_item.grid(row=0, column=1)
        entry_unit.grid(row=1, column=1)
        entry_price.grid(row=2, column=1)
        entry_stock.grid(row=3, column=1)

        entry_item.bind('<Tab>', event_search, True)
        entry_item.bind('<Right>', event_search, True)
        entry_item.bind('<Return>', event_search, True)
        entry_item.bind('<Return>', event_item, True)
        entry_unit.bind('<Return>', event_unit)
        entry_price.bind('<Return>', event_price)
        entry_stock.bind('<Return>', event_stock)
        # </editor-fold>
        # </editor-fold>

        # <editor-fold desc="Data Mid Top Left Frame">
        d_mid_top_left = Frame(d_mid_top, padding=(0, 0, 0, 10))
        d_mid_top_left.pack(side=LEFT, fill=Y)

        # <editor-fold desc="Data Treeview">
        cols = {
            '#0': ['No', 70, 'w'],
            'item': ['Item', 270, 'w'],
            'unit': ['Unit', 100, 'center'],
            'price': ['Price', 150, 'e'],
            'stock': ['Stock', 150, 'e'],
        }
        tv_data = self.treeviews(d_mid_top_left, cols, 20)
        tv_data.bind('<<TreeviewSelect>>', tv_select)
        # </editor-fold>
        # </editor-fold>
        # </editor-fold>

        # <editor-fold desc="Data Buttons">
        btn_texts = ['Add New', 'Display', 'Clear',
                     'Delete', 'Search', 'Update', 'Exit']
        buttons = self.buttons(d_mid_frame, btn_texts)
        buttons[0]['command'] = add_new
        buttons[1]['command'] = display_data
        buttons[2]['command'] = clear_entries
        buttons[3]['command'] = delete_data
        buttons[4]['command'] = search
        buttons[5]['command'] = update
        # </editor-fold>
        # </editor-fold>

        # <editor-fold desc="Data Footer">
        self.footer(data_tab)
        # </editor-fold>
        # </editor-fold>

    def transac_win(self):

        # # <editor-fold desc="Transactions Functions">
        def add_item():
            data_keys = self.sort_data_dict()
            i, u, q, t, d = entry_item.get(), entry_unit.get(
            ), entry_qty.get(), entry_tot.get(), entry_discnt.get()
            if len(i) != 0 and len(u) != 0:
                x = backend.search_data(i, u)
                x = x[0][0]
                key = f'{i}|{u}'
                if key in data_keys:
                    if len(q) != 0 and len(t) != 0:
                        tot = float(t)
                        qty = float(q)
                        if self.data_dict[key][1] >= qty:
                            s = self.data_dict[key][1] - qty
                            sub = self.data_dict[key][0] * qty
                            if key in self.items:
                                nq = self.items[key][0] + qty
                                tot = tot + self.items[key][3]
                                sub = sub + self.items[key][1]
                            else:
                                nq = qty
                            d = sub - tot
                            self.items[key] = [
                                round(nq, 2), sub, round(d), tot]
                            backend.update_data(
                                x, i, u, self.data_dict[key][0], s)
                        else:
                            messagebox.showwarning(
                                'Warning', f'Not enough stock for {i} in {u}')
                    else:
                        messagebox.showwarning(
                            'Warning', '"Quantity" and "Total" should not be empty')
                else:
                    messagebox.showwarning(
                        'Warning', f'{i} in {u} is not in the database')
            else:
                messagebox.showwarning(
                    'Warning', '"Item" and "Unit" should not be empty')
            display_items()
            display_all()

        def search_item():
            i, u = entry_item.get(), entry_unit.get()
            if len(i) == 0 and len(u) == 0:
                display_all()
            else:
                searched_data = backend.search_data(i, u)

                data = []
                for i in searched_data:
                    d = [i[1], i[2], i[3], i[4]]
                    data.append(d)

                self.clear_tv(tv_left)
                self.display(tv_left, data)

        def clear_item():
            entry_item.delete(0, END)
            entry_unit.delete(0, END)
            entry_qty.delete(0, END)
            entry_tot.delete(0, END)
            entry_discnt.delete(0, END)
            entry_item.focus_set()

        def delete_item():
            i, u, q = entry_item.get(), entry_unit.get(), entry_qty.get()
            if len(i) != 0 and len(u) != 0 and len(q) != 0:
                k = f'{i}|{u}'
                x = backend.search_data(i, u)
                x = x[0]
                if self.items[k][0] >= float(q):
                    nq = self.items[k][0] - float(q)
                    p = self.items[k][3] / self.items[k][0]
                    d = float(q) * p
                    nt = self.items[k][3] - d
                    if nq == 0:
                        self.items.pop(k)
                    else:
                        sub = self.items[k][1] - x[3] * float(q)
                        self.items[k] = [nq, sub, sub - nt, nt]
                    s = x[4] + float(q)
                    backend.update_data(x[0], i, u, x[3], s)
                else:
                    messagebox.showwarning(
                        'Warning', 'Invalid "Quantity" value')
            display_items()
            display_all()

        def cancel():
            if self.items:
                a = messagebox.askyesno(
                    'Cancel', 'Confirm if you want to cancel the transaction')
                if a > 0:
                    data = self.items
                    for k, v in data.items():
                        i, u = k.split('|')
                        x = backend.search_data(i, u)
                        x = x[0]
                        s = x[4] + v[0]
                        backend.update_data(x[0], i, u, x[3], s)
                    self.items.clear()
                    clear_item()
                display_all()
                display_items()

        def pay():
            if self.items:
                a = messagebox.askyesno(
                    'Payment', 'Confirm and accept the payment')
                if a > 0:
                    h, m = str(datetime.now())[11:-10].split(':')
                    if int(h) < 12:
                        ap = 'AM'
                    elif int(h) == 0:
                        ap = 'AM'
                        h = 12
                    else:
                        ap = 'PM'
                        h = int(h) - 12
                    time = f'{h}:{m} {ap}'
                    data = self.items
                    to_file = []
                    for k, v in data.items():
                        i, u = k.split('|')
                        to_file.append([time, v[0], u, i, v[1], v[2], v[3]])
                    save_file(to_file)
                    self.items.clear()
                    clear_item()
                    display_items()
                    entry_cash.delete(0, END)
                    lbl_change_change.config(text='Php 0')

        def save_file(data):
            sheet_name = str(date.today())
            try:
                wb = load_workbook('MalabuyocSales.xlsx')
            except FileNotFoundError:
                wb = Workbook()
            try:
                ws = wb[sheet_name]
            except KeyError:
                ws = wb.create_sheet(sheet_name)
                ws.append(['Time', 'Qty', 'Unit', 'Item', 'Sub',
                           'Disc', 'Total', 'Total Sales', '=SUM(G2:G500)'])
            for row in data:
                ws.append(row)
            wb.save('MalabuyocSales.xlsx')

        def display_items():
            data = self.items
            for i in range(1, len(data) + 10):
                if tv_right.exists(i):
                    tv_right.delete(i)
            id = 1
            self.due = 0
            for k, v in data.items():
                i, u = k.split('|')
                tv_right.insert('', END, id, text=id, values=(
                    v[0], u, i, v[1], v[2], v[3]))
                self.due = self.due + v[3]
                id += 1
            if tv_right.exists(1):
                tv_right.see(id - 1)
            lbl_items_items.config(text=len(data))
            lbl_due.config(text=f'Php {round(self.due, 2)}')

        def display_all():

            keys = self.sort_data_dict()
            data = []
            for k in keys:
                i, u = k.split('|')
                p, s = self.data_dict[k][0], self.data_dict[k][1]
                d = [i, u, p, s]
                data.append(d)

            self.clear_tv(tv_left)
            self.display(tv_left, data)

        def tv_left_select(event):
            f = tv_left.focus()
            i = tv_left.item(f)
            x = i['values']
            entry_item.delete(0, END)
            entry_item.insert(END, x[0])
            entry_unit.delete(0, END)
            entry_unit.insert(END, x[1])
            entry_qty.delete(0, END)
            entry_tot.delete(0, END)
            entry_discnt.delete(0, END)

        def tv_right_select(event):
            f = tv_right.focus()
            i = tv_right.item(f)
            x = i['values']
            entry_item.delete(0, END)
            entry_item.insert(END, x[2])
            entry_unit.delete(0, END)
            entry_unit.insert(END, x[1])
            entry_qty.delete(0, END)
            entry_qty.insert(END, x[0])
            entry_tot.delete(0, END)
            entry_tot.insert(END, x[3])
            entry_discnt.delete(0, END)

        def calc_change(event):
            cash = entry_cash.get()
            if len(cash) != 0:
                c = float(cash) - self.due
                lbl_change_change.config(text=f'Php {c}')

        def event_search(event):
            search_item()

        def event_item(event):
            entry_unit.focus_set()

        def event_unit(event):
            entry_qty.focus_set()

        def event_qty(event):
            if len(entry_qty.get()) == 0:
                entry_tot.focus_set()
            else:
                i, u = entry_item.get(), entry_unit.get()
                if len(i) != 0 and len(u) != 0:
                    x = backend.search_data(i, u)
                    x = x[0]
                    tot = float(entry_qty.get()) * x[3]
                    entry_tot.delete(0, END)
                    entry_tot.insert(END, tot)

        def event_tot(event):
            if len(entry_tot.get()) == 0:
                add_item()
            else:
                i, u = entry_item.get(), entry_unit.get()
                if len(i) != 0 and len(u) != 0:
                    x = backend.search_data(i, u)
                    x = x[0]
                    qty = round((float(entry_tot.get()) / x[3]), 2)
                    entry_qty.delete(0, END)
                    entry_qty.insert(END, qty)

        def event_dis_price(event):
            if len(entry_discnt.get()) != 0:
                dp = float(entry_discnt.get())
                q = float(entry_qty.get())
                tot = dp * q
                entry_tot.delete(0, END)
                entry_tot.insert(END, tot)

        def event_add(event):
            add_item()
        # </editor-fold>

        # <editor-fold desc="Transactions Frame">
        transac_tab = Frame(self.tabs, padding=(
            50, 0, 50, 0), style='Main.TFrame')
        transac_tab.pack()

        # Add Transactions Tab
        self.tabs.add(transac_tab, text='Transaction')

        # Transactions Title
        self.title(transac_tab, self.trans_title)

        # <editor-fold desc="Transactions Mid Frame">
        t_mid = Frame(transac_tab, relief=SUNKEN,
                      borderwidth=5, padding=(30, 10, 30, 10))
        t_mid.pack(fill=X)

        # <editor-fold desc="Transactions Mid Top Frame">
        t_mid_top = Frame(t_mid)
        t_mid_top.pack(side=TOP, fill=X)

        # <editor-fold desc="Transactions Left Frame">
        t_left = Frame(t_mid_top, width=500)
        t_left.pack(side=LEFT, fill=Y)

        # <editor-fold desc="Transactions Left Top Frame">
        t_left_top = Frame(t_left)
        t_left_top.pack(side=TOP, fill=X)

        # <editor-fold desc="Transactions Left Treeview">
        cols_left = {
            '#0': ['No', 55, 'w'],
            'item': ['Item', 200, 'w'],
            'unit': ['Unit', 70, 'center'],
            'price': ['Price', 80, 'e'],
            'stock': ['Stock', 80, 'e'],
        }
        tv_left = self.treeviews(t_left_top, cols_left, 10)
        tv_left.bind('<<TreeviewSelect>>', tv_left_select)
        # </editor-fold>
        # </editor-fold>

        # <editor-fold desc="Transactions Left Bottom Frame">
        t_left_bot = Frame(t_left)
        t_left_bot.pack(side=BOTTOM, fill=X, pady=6)

        # <editor-fold desc="Transaction Left Labels">
        lbl_item = Label(t_left_bot, text='Item')
        lbl_unit = Label(t_left_bot, text='Unit')
        lbl_qty = Label(t_left_bot, text='Quantity')
        lbl_tot = Label(t_left_bot, text='Set Amount')
        lbl_discnt = Label(t_left_bot, text='Set Price')

        lbl_item.grid(row=0, column=0, sticky=E)
        lbl_unit.grid(row=1, column=0, sticky=E)
        lbl_qty.grid(row=2, column=0, sticky=E)
        lbl_tot.grid(row=3, column=0, sticky=E)
        lbl_discnt.grid(row=4, column=0, sticky=E)
        # </editor-fold>

        # <editor-fold desc="Transactions Left Entries">
        entry_item = AutocompleteEntry(
            self.data_list, master=t_left_bot, font='arial 18 bold', foreground='green')
        entry_unit = Entry(t_left_bot, font='arial 18 bold',
                           foreground='green')
        entry_qty = Entry(t_left_bot, font='arial 18 bold', foreground='green')
        entry_tot = Entry(t_left_bot, font='arial 18 bold', foreground='green')
        entry_discnt = Entry(
            t_left_bot, font='arial 18 bold', foreground='green')

        entry_item.grid(row=0, column=1)
        entry_unit.grid(row=1, column=1)
        entry_qty.grid(row=2, column=1)
        entry_tot.grid(row=3, column=1)
        entry_discnt.grid(row=4, column=1)

        entry_item.focus_set()
        entry_item.bind('<Return>', event_item, True)
        entry_item.bind('<Return>', event_search, True)
        entry_item.bind('<Tab>', event_search, True)
        entry_item.bind('<Right>', event_search, True)
        entry_unit.bind('<Return>', event_unit)
        entry_qty.bind('<Return>', event_qty)
        entry_qty.bind('<Control-Return>', event_add, True)
        entry_tot.bind('<Return>', event_tot)
        entry_tot.bind('<Control-Return>', event_add, True)
        entry_discnt.bind('<Return>', event_dis_price)
        entry_discnt.bind('<Control-Return>', event_add, True)
        # </editor-fold>
        # </editor-fold>
        # </editor-fold>

        # <editor-fold desc="Transactions Right Frame">
        t_right = Frame(t_mid_top)
        t_right.pack(side=RIGHT, fill=Y)

        # <editor-fold desc="Transactions Right Top Frame">
        t_right_top = Frame(t_right)
        t_right_top.pack(side=TOP, fill=X)
        t_title_right = Label(
            t_right_top, text='Transaction Summary', anchor=CENTER, font='arial 24 bold')
        t_title_right.pack()
        # </editor-fold>

        # <editor-fold desc="Transactions Right Mid Frame">
        t_right_mid = Frame(t_right)
        t_right_mid.pack(fill=X)

        # <editor-fold desc="Transactions Right Treeview">
        cols_right = {
            '#0': ['No', 50, 'w'],
            'qty': ['Qty', 55, 'e'],
            'unit': ['Unit', 70, 'center'],
            'item': ['Item', 200, 'w'],
            'sub': ['Sub', 80, 'e'],
            'disc': ['Disc', 75, 'e'],
            'total': ['Total', 80, 'e']
        }
        tv_right = self.treeviews(t_right_mid, cols_right, 10)
        tv_right.bind('<<TreeviewSelect>>', tv_right_select)
        # </editor-fold>
        # </editor-fold>

        # <editor-fold desc="Transactions Right Bottom Frame">
        t_right_bot_dummy = Frame(t_right)
        t_right_bot_dummy.pack(side=BOTTOM, fill=X)
        t_right_bot = Frame(t_right_bot_dummy)
        t_right_bot.pack()

        # <editor-fold desc="Transactions Right Labels And Entries">
        lbl_items = Label(t_right_bot, text='Items')
        lbl_items_items = Label(t_right_bot, text='0', foreground='red')
        lbl_amt = Label(t_right_bot, text='Amount Due')
        lbl_due = Label(t_right_bot, text='Php 0', foreground='red')
        lbl_cash = Label(t_right_bot, text='Cash')
        entry_cash = Entry(t_right_bot, font='arial 18 bold', foreground='red')
        lbl_change = Label(t_right_bot, text='Change')
        lbl_change_change = Label(t_right_bot, text='Php 0', foreground='red')

        lbl_items.grid(row=2, column=0, sticky=E)
        lbl_items_items.grid(row=2, column=1, sticky=W)
        lbl_amt.grid(row=3, column=0, sticky=E)
        lbl_due.grid(row=3, column=1, sticky=W)
        lbl_cash.grid(row=4, column=0, sticky=E)
        entry_cash.grid(row=4, column=1, sticky=W)
        lbl_change.grid(row=5, column=0, sticky=E)
        lbl_change_change.grid(row=5, column=1, sticky=W)

        entry_cash.bind('<Return>', calc_change)
        # </editor-fold>
        # </editor-fold>
        # </editor-fold>
        # </editor-fold>

        # <editor-fold desc="Transactions Buttons">
        btn_texts = ['Add Item', 'Search', 'Clear',
                     'Delete', 'Cancel', 'Pay', 'Exit']
        buttons = self.buttons(t_mid, btn_texts)
        buttons[0]['command'] = add_item
        buttons[1]['command'] = search_item
        buttons[2]['command'] = clear_item
        buttons[3]['command'] = delete_item
        buttons[4]['command'] = cancel
        buttons[5]['command'] = pay
        # </editor-fold>
        # </editor-fold>

        # <editor-fold desc="Transactions Footer">
        self.footer(transac_tab)
        # </editor-fold>
        # </editor-fold>

    def sales_win(self):

        # <editor-fold desc="Sales Functions">
        def get_sales_data():
            year = combo_y.get()
            month = combo_m.get()
            day = combo_d.get()
            if len(year) == 0 and len(month) == 0 and len(day) == 0:
                sheet_name = str(date.today())
            elif len(year) != 0 and (len(month) == 0 or len(day) == 0):
                sheet_name = str(date.today())
            elif len(year) == 0 and (len(month) != 0 and len(day) != 0):
                sheet_name = f'2019-{month}-{day}'
            else:
                month = months_inverse[month]
                if len(day) == 1:
                    day = f'0{day}'
                sheet_name = f'{year}-{month}-{day}'

            yy, mm, dd = sheet_name.split('-')
            mmm = months[mm]
            s_title.config(text=f'{mmm} {dd}, {yy}')
            wb = load_workbook('MalabuyocSales.xlsx', data_only=True)
            data = []
            try:
                ws = wb[sheet_name]
            except KeyError:
                messagebox.showwarning(
                    'Warning', f'The sales sheet in {yy}-{mm}-{dd} does not exist')
                return data
            for row in ws.iter_rows():
                to_data = [row[0].value, row[1].value, row[2].value,
                           row[3].value, row[4].value, row[5].value, row[6].value]
                data.append(to_data)
            return data[1:]

        def display_sales():
            data = get_sales_data()
            for i in range(1, len(data) + 10):
                if tv_sale.exists(i):
                    tv_sale.delete(i)
            id = 1
            ts = 0
            for row in data:
                tv_sale.insert('', END, id, text=id, values=(
                    row[0], row[1], row[2], row[3], row[4], row[5], row[6]))
                id += 1
                ts = ts + row[6]
            if tv_sale.exists(1):
                tv_sale.see(id - 1)
            lbl_tot_sales.config(text=f'Php {ts}')
        # </editor-fold>

        # <editor-fold desc="Sales Frame">
        s_tab = Frame(self.tabs, style='Main.TFrame', padding=(50, 0, 50, 0))
        s_tab.pack()

        # Add Sales Tab
        self.tabs.add(s_tab, text='Sales')

        # Sales Title
        s_title = self.title(s_tab, self.sales_title)

        # <editor-fold desc="Sales Mid Frame">
        s_mid = Frame(s_tab, relief=SUNKEN, borderwidth=5,
                      padding=(30, 10, 30, 10))
        s_mid.pack(fill=X)

        # <editor-fold desc="Sales Mid Top Frame">
        s_mid_top = Frame(s_mid, padding=(0, 0, 0, 10))
        s_mid_top.pack(side=TOP, fill=X)

        # <editor-fold desc="Sales Mid Top Left Frame">
        s_mid_top_left = Frame(s_mid_top)
        s_mid_top_left.pack(side=LEFT, fill=Y)

        # <editor-fold desc="Sales Treeview">
        cols = {
            '#0': ['No', 60, 'w'],
            'time': ['Time', 100, 'center'],
            'qty': ['Qty', 80, 'e'],
            'unit': ['Unit', 80, 'center'],
            'item': ['Item', 200, 'w'],
            'sub': ['Sub', 100, 'e'],
            'disc': ['Disc', 90, 'e'],
            'total': ['Total', 100, 'e']
        }
        tv_sale = self.treeviews(s_mid_top_left, cols, 20)
        # </editor-fold>
        # </editor-fold>

        # <editor-fold desc="Sales Mid Top Right Frame">
        s_mid_top_right = Frame(s_mid_top, padding=(0, 20, 20, 0))
        s_mid_top_right.pack(side=RIGHT, fill=Y)

        # <editor-fold desc="Sales Right Labels">
        lbl_t = Label(s_mid_top_right, text='Display sales of',
                      font='arial 20 bold')
        lbl_m = Label(s_mid_top_right, text='Month', font='arial 16')
        lbl_d = Label(s_mid_top_right, text='Day', font='arial 16')
        lbl_y = Label(s_mid_top_right, text='Year', font='arial 16')
        lbl = Label(s_mid_top_right, text='', font='arial 20 bold')
        lbl_tot_sale = Label(s_mid_top_right, text='Total Sales')
        lbl_tot_sales = Label(s_mid_top_right, text='Php 0.0',
                              foreground='red', font='arial 30 bold')

        lbl_t.grid(row=0, columnspan=2)
        lbl_m.grid(row=1, column=0, sticky=E)
        lbl_d.grid(row=2, column=0, sticky=E)
        lbl_y.grid(row=3, column=0, sticky=E)
        lbl.grid(row=4, columnspan=2)
        lbl_tot_sale.grid(row=5, columnspan=2)
        lbl_tot_sales.grid(row=6, columnspan=2)
        # </editor-fold>

        # <editor-fold desc="Sales Comboboxes">
        combo_m = Combobox(s_mid_top_right, values=mos,
                           font='arial 16', width=15)
        combo_d = Combobox(s_mid_top_right, values=days,
                           font='arial 16', width=15)
        combo_y = Combobox(s_mid_top_right, values=yrs,
                           font='arial 16', width=15)

        combo_m.grid(row=1, column=1)
        combo_d.grid(row=2, column=1)
        combo_y.grid(row=3, column=1)
        # </editor-fold>
        # </editor-fold>
        # </editor-fold>

        # <editor-fold desc="Sales Buttons">
        btn_texts = ['Update', 'Exit']
        buttons = self.buttons(s_mid, btn_texts)
        buttons[0]['command'] = display_sales
        # </editor-fold>
        # </editor-fold>

        # <editor-fold desc="Footer">
        self.footer(s_tab)
        # </editor-fold>
        # </editor-fold>

    # <editor-fold desc="Methods">
    def title(self, parent, title):

        title_frame = Frame(parent, style='Main.TFrame')
        title_frame.pack(side=TOP, fill=X)

        title = Label(title_frame, background='cadet blue',
                      foreground='orange', text=title, font='arial 40 bold')
        title_info = Label(title_frame, background='cadet blue', foreground='white',
                           text='Nug-as, Alcoy, Cebu & Poblacion, Malabuyoc, Cebu | Contact: 09179720929 | Email: rrychemae.medel@yahoo.com', font='arial 14 bold')

        title.pack()
        title_info.pack()

        return title

    def treeviews(self, parent, cols, height):

        tv_frame = Frame(parent)
        tv_frame.pack()

        cols_k = []
        for k, _ in cols.items():
            cols_k.append(k)

        scroll = Scrollbar(tv_frame)
        scroll.grid(row=0, column=1, sticky=NS)

        tv = Treeview(tv_frame, height=height, yscrollcommand=scroll.set)
        tv.grid(row=0, column=0)

        tv['columns'] = cols_k[1:]

        for k, v in cols.items():
            tv.column(k, width=v[1], anchor=v[2])
            tv.heading(k, text=v[0])

        scroll['command'] = tv.yview

        return tv

    def buttons(self, parent, btn_texts):

        bot_frame = Frame(parent)
        bot_frame.pack(side=BOTTOM, fill=X)
        btn_frame = Frame(bot_frame)
        btn_frame.pack()

        buttons = []
        for i, t in enumerate(btn_texts):
            btn = Button(btn_frame, text=t)
            btn.grid(row=0, column=i)
            buttons.append(btn)

        buttons[-1]['command'] = self.stop

        return buttons

    def footer(self, parent):

        footer = Frame(parent, style='Main.TFrame')
        footer.pack(side=BOTTOM, fill=X)

        footer_label = Label(footer, background='cadet blue',
                             text='© jiMcaN | jimcan49@yahoo.com', foreground='brown', font='arial 14 normal')
        footer_label.pack()

    def update_data_list(self):

        self.data_list.clear()

        data = backend.view_data()

        for d in data:
            if d[1] not in self.data_list:
                self.data_list.append(d[1])

    def sort_data_dict(self):

        self.data_dict.clear()

        data = backend.view_data()
        for d in data:
            k = f'{d[1]}|{d[2]}'
            self.data_dict[k] = [d[3], d[4]]

        return sorted(self.data_dict)

    def clear_tv(self, tv):
        tv_children = tv.get_children()
        for i in range(1, len(tv_children) + 1):
            tv.delete(i)

    def display(self, tv, data):

        id = 1
        for row in data:
            tv.insert('', END, id, text=id, values=(
                row[0], row[1], row[2], row[3]))
            id += 1

        if id > 1:
            tv.see(id - 1)

    def run(self):

        self.root.mainloop()

    def stop(self):

        x = messagebox.askyesno(
            'Exit Application', 'Confirm if you want to exit')

        if x > 0:
            if self.items:
                data = self.items
                for k, v in data.items():
                    i, u = k.split('|')
                    x = backend.search_data(i, u)
                    x = x[0]
                    s = x[4] + v[0]
                    backend.update_data(x[0], i, u, x[3], s)

            self.root.destroy()
    # </editor-fold>


if __name__ == '__main__':
    Main().run()
